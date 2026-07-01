import os
import time
import asyncio
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_FILE = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "api_key_metrics.xlsx")
log_queue = asyncio.Queue()

def initialize_excel():
    """Create the Excel file with premium styling if it does not exist."""
    if os.path.exists(EXCEL_FILE):
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "API Key Metrics"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers definition
    headers = [
        "Timestamp", "Session ID", "Action", "Provider", 
        "Key Label", "Model", "Task Type", "Status", 
        "Tokens Used", "Latency (s)", "Error / Logs"
    ]
    
    ws.append(headers)
    
    # Styles for the header row
    header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Deep Premium Navy Blue
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="medium", color="1A365D")
    )
    
    # Style header cells
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Set premium row height for header
    ws.row_dimensions[1].height = 28
    
    # Set default column widths for premium reading spacing
    col_widths = {
        "A": 22, # Timestamp
        "B": 36, # Session ID
        "C": 15, # Action
        "D": 15, # Provider
        "E": 25, # Key Label
        "F": 25, # Model
        "G": 15, # Task Type
        "H": 12, # Status
        "I": 15, # Tokens Used
        "J": 12, # Latency
        "K": 50  # Error / Logs
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
        
    wb.save(EXCEL_FILE)

def append_row_to_excel(row_data):
    """Deterministically appends a row of metric log data with professional cell borders and formatting."""
    initialize_excel()
    
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Append data row
        ws.append(row_data)
        row_idx = ws.max_row
        
        # Styles for data row
        font_style = Font(name="Segoe UI", size=10)
        border_style = Border(
            left=Side(style="thin", color="E0E0E0"),
            right=Side(style="thin", color="E0E0E0"),
            top=Side(style="thin", color="E0E0E0"),
            bottom=Side(style="thin", color="E0E0E0")
        )
        
        # Alternating row background colors (zebra striping for readability)
        bg_color = "F7FAFC" if row_idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # Custom status styles
        status = str(row_data[7]).upper()
        if status == "SUCCESS":
            status_font = Font(name="Segoe UI", size=10, bold=True, color="2F855A") # Forest Green
            status_fill = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid") # Mint
        elif status in ("FAILED", "QUARANTINED", "FAILURE"):
            status_font = Font(name="Segoe UI", size=10, bold=True, color="C53030") # Deep Red
            status_fill = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid") # Soft Red
        else:
            status_font = font_style
            status_fill = row_fill

        ws.row_dimensions[row_idx].height = 20
        
        # Apply styles cell by cell
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_style
            
            # Custom styling for status cell
            if col_idx == 8: # Status column
                cell.font = status_font
                cell.fill = status_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = font_style
                cell.fill = row_fill
                # Center-align specific columns
                if col_idx in (1, 3, 4, 7, 9, 10):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[KeyOptimus Excel Logger Error] Failed to write to Excel: {e}")

async def excel_writer_worker():
    """Background task that drains the log queue and writes entries to Excel sequentially to avoid file locks."""
    print("[KeyOptimus Excel Logger] Background writer worker started.")
    while True:
        try:
            row_data = await log_queue.get()
            # Run blocking file operations in an executor to avoid locking the event loop
            await asyncio.to_thread(append_row_to_excel, row_data)
            log_queue.task_done()
        except asyncio.CancelledError:
            break
        except Exception as e:
            print(f"[KeyOptimus Worker Loop Error] {e}")
            await asyncio.sleep(1)

def queue_log(session_id, action, provider, key_label, model, task_type, status, tokens_used=0, latency=0.0, error_msg=""):
    """Thread-safe and async-safe interface to submit a log entry to the queue."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row_data = [
        timestamp,
        session_id or "N/A",
        action or "ROUTE",
        provider or "N/A",
        key_label or "N/A",
        model or "N/A",
        task_type or "text",
        status or "SUCCESS",
        tokens_used,
        round(latency, 3),
        error_msg or ""
    ]
    # Queue it from running event loop
    try:
        loop = asyncio.get_running_loop()
        loop.call_soon_threadsafe(log_queue.put_nowait, row_data)
    except RuntimeError:
        # Fallback if no loop is running (e.g., during startup checks or test scripts)
        append_row_to_excel(row_data)
